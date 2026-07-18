"""
Export einer Rechnung als .xlsx - im selben Layout wie die bisherige
handgepflegte Excel-Vorlage, aber mit automatisch berechneten Formeln.
"""
import openpyxl
from openpyxl.styles import Font
from app.invoice import berechne_rechnung, KOPIEN_GRENZE, KOPIEN_SATZ_BIS_GRENZE, KOPIEN_SATZ_AB_GRENZE


def exportiere_rechnung_xlsx(
    pfad: str,
    fall: dict,
    rechnung: dict,
    zeitposten: list,
    einstellungen: dict,
):
    ergebnis = berechne_rechnung(
        zeitposten=zeitposten,
        stundensatz=rechnung["stundensatz"],
        km=rechnung["km"],
        km_satz=rechnung["km_satz"],
        porto=rechnung["porto"],
        telefon=rechnung["telefon"],
        zeichen_anzahl=rechnung["zeichen_anzahl"],
        schreibgebuehr_satz=rechnung["schreibgebuehr_satz"],
        kopien_seiten=rechnung["kopien_seiten"],
        mwst_satz=rechnung["mwst_satz"],
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rechnung"
    bold = Font(bold=True)

    row = 1
    ws.cell(row=row, column=1, value=rechnung.get("datum", "")); row += 1
    ws.cell(row=row, column=1, value=f"An: {fall.get('empfaenger', fall.get('gericht',''))}"); row += 1
    ws.cell(row=row, column=1, value="Kostennote: Familienpsychologisches Sachverständigengutachten"); row += 1
    ws.cell(row=row, column=1, value=f"Gericht: {fall.get('gericht','')}/ {fall.get('abteilung','')}"); row += 1
    ws.cell(row=row, column=1, value=f"In Sachen: {fall.get('in_sachen','')}"); row += 1
    ws.cell(row=row, column=1, value=f"Aktenzeichen: {fall.get('aktenzeichen','')}"); row += 1
    ws.cell(row=row, column=1, value=f"Rechnungsnummer: {rechnung.get('rechnungsnummer','')}"); row += 1
    row += 1

    ws.cell(row=row, column=1, value="1. Zeitaufwand").font = bold
    row += 2
    zeit_start = row
    for posten in zeitposten:
        ws.cell(row=row, column=1, value=posten.get("bezeichnung", ""))
        ws.cell(row=row, column=6, value=int(posten.get("minuten", 0) or 0))
        row += 1
    zeit_ende = row - 1
    row += 1
    ws.cell(row=row, column=1, value="Minuten")
    ws.cell(row=row, column=6, value=f"=SUM(F{zeit_start}:F{zeit_ende})")
    row += 1
    ws.cell(row=row, column=1, value="Stunden")
    ws.cell(row=row, column=6, value=f"=F{row-1}/60")
    row += 2
    ws.cell(
        row=row, column=1,
        value=f"aufgerundet {ergebnis.stunden_aufgerundet:.2f} Stunden à {ergebnis.stundensatz:.2f} €",
    )
    ws.cell(row=row, column=6, value=ergebnis.summe_zeitaufwand)
    row += 2
    ws.cell(row=row, column=1, value="Summe 1").font = bold
    ws.cell(row=row, column=6, value=ergebnis.summe_zeitaufwand)
    summe1_row = row
    row += 2

    ws.cell(row=row, column=1, value="2. Aufwendungen").font = bold
    row += 2
    aufw_start = row
    ws.cell(row=row, column=1, value="Reisekosten")
    ws.cell(row=row, column=2, value="km")
    ws.cell(row=row, column=3, value=rechnung["km"])
    ws.cell(row=row, column=4, value="à")
    ws.cell(row=row, column=5, value=rechnung["km_satz"])
    ws.cell(row=row, column=6, value=ergebnis.reisekosten)
    row += 1
    ws.cell(row=row, column=1, value="Porto")
    ws.cell(row=row, column=6, value=ergebnis.porto)
    row += 1
    ws.cell(row=row, column=1, value="Telefon")
    ws.cell(row=row, column=6, value=ergebnis.telefon)
    row += 1
    ws.cell(row=row, column=1, value="Schreibgebühr")
    ws.cell(row=row, column=2, value="Zeichen")
    ws.cell(row=row, column=3, value=rechnung["zeichen_anzahl"])
    ws.cell(row=row, column=4, value=f"je angef. 1000 x {rechnung['schreibgebuehr_satz']} €")
    ws.cell(row=row, column=6, value=ergebnis.schreibgebuehr)
    row += 1
    seiten = rechnung["kopien_seiten"] or 0
    seiten_bis = min(seiten, KOPIEN_GRENZE)
    seiten_ueber = max(0, seiten - KOPIEN_GRENZE)
    ws.cell(row=row, column=1, value="Kopien GA")
    ws.cell(row=row, column=2, value="Seiten")
    ws.cell(row=row, column=3, value=seiten_bis)
    ws.cell(row=row, column=4, value="à")
    ws.cell(row=row, column=5, value=KOPIEN_SATZ_BIS_GRENZE)
    ws.cell(row=row, column=6, value=round(seiten_bis * KOPIEN_SATZ_BIS_GRENZE, 2))
    row += 1
    if seiten_ueber > 0:
        ws.cell(row=row, column=2, value="Seiten")
        ws.cell(row=row, column=3, value=seiten_ueber)
        ws.cell(row=row, column=4, value="à")
        ws.cell(row=row, column=5, value=KOPIEN_SATZ_AB_GRENZE)
        ws.cell(row=row, column=6, value=round(seiten_ueber * KOPIEN_SATZ_AB_GRENZE, 2))
        row += 1
    aufw_ende = row - 1
    row += 1
    ws.cell(row=row, column=1, value="Summe 2").font = bold
    ws.cell(row=row, column=6, value=f"=SUM(F{aufw_start}:F{aufw_ende})")
    summe2_row = row
    row += 2

    ws.cell(row=row, column=1, value="Summe 1+2").font = bold
    ws.cell(row=row, column=6, value=f"=F{summe1_row}+F{summe2_row}")
    summe12_row = row
    row += 1
    ws.cell(row=row, column=1, value=f"Mehrwertsteuer {ergebnis.mwst_satz:.0f}%")
    ws.cell(row=row, column=6, value=f"=F{summe12_row}*{ergebnis.mwst_satz:.0f}%")
    mwst_row = row
    row += 1
    ws.cell(row=row, column=1, value="Gesamtsumme").font = bold
    ws.cell(row=row, column=6, value=f"=F{summe12_row}+F{mwst_row}")
    row += 2

    ws.cell(row=row, column=1, value=f" {einstellungen.get('bank','')}; IBAN: {einstellungen.get('iban','')}")
    row += 1
    ws.cell(row=row, column=1, value=f"Kontoinhaberin: {einstellungen.get('kontoinhaberin','')}")
    row += 1
    ws.cell(row=row, column=1, value=f"Finanzamt {einstellungen.get('finanzamt','')}; Steuernummer: {einstellungen.get('steuernummer','')}; USt.IdNr.: {einstellungen.get('ust_idnr','')}")
    row += 1
    ws.cell(row=row, column=1, value=f"Steuer-ID: {einstellungen.get('steuer_id','')}")

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["F"].width = 12

    wb.save(pfad)
    return ergebnis
